"""Testes da resolução de abas renomeadas (``infra.abas``).

Cobre o caso real que quebrou a atualização: a base de postos chegou com a aba
"Dados" renomeada para "Planilha1" (conteúdo idêntico) e o pipeline parou no 1º
passo. A aba precisa ser reencontrada pelo cabeçalho, sem afrouxar a ponto de
aceitar uma aba qualquer.
"""

import unittest

import openpyxl

from app_oficinas import config
from app_oficinas.errors import AbaNaoEncontrada
from app_oficinas.infra import abas

CAB_POSTOS = ["Frete", "MP", "Oficinas", "Data Efetivos", "QTD Efetivos",
              "Data Trabalhados", "QTD Trabalhados", "Contratação",
              "Demissão", "Semana"]

ARQ = config.ABSENTEISMO.arquivo
ABA = config.ABSENTEISMO.aba


def _wb(abas_e_cabecalhos: dict[str, list]) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for titulo, cabecalho in abas_e_cabecalhos.items():
        ws = wb.create_sheet(titulo)
        ws.append(cabecalho)
    return wb


class TestResolver(unittest.TestCase):
    def test_nome_exato_vence(self):
        wb = _wb({"Dados": CAB_POSTOS, "Planilha1": CAB_POSTOS})
        self.assertEqual(abas.resolver(wb, ABA, ARQ), "Dados")

    def test_ignora_acento_e_caixa(self):
        wb = _wb({"DADOS": ["qualquer"]})
        self.assertEqual(abas.resolver(wb, ABA, ARQ), "DADOS")

    def test_encontra_pelo_cabecalho_quando_aba_foi_renomeada(self):
        # O caso real: "Dados" virou "Planilha1" numa reexportação do Excel.
        wb = _wb({"Planilha1": CAB_POSTOS, "Planilha2": []})
        self.assertEqual(abas.resolver(wb, ABA, ARQ), "Planilha1")

    def test_erro_quando_nenhuma_aba_tem_o_cabecalho(self):
        wb = _wb({"Planilha1": ["Nome", "Valor"], "Resumo": ["A"]})
        with self.assertRaises(AbaNaoEncontrada) as ctx:
            abas.resolver(wb, ABA, ARQ)
        # A mensagem precisa listar o que existe, para o usuário se orientar.
        self.assertIn("Planilha1", str(ctx.exception))

    def test_nao_escolhe_no_escuro_quando_duas_abas_casam(self):
        # Ambiguidade é erro: ler a aba errada publicaria números inventados.
        wb = _wb({"Base A": CAB_POSTOS, "Base B": CAB_POSTOS})
        with self.assertRaises(AbaNaoEncontrada):
            abas.resolver(wb, ABA, ARQ)

    def test_cabecalho_fora_da_linha_1(self):
        # RESUMO/DEFEITOS do "Indicador geral" têm cabeçalho na linha 2.
        arq, aba = config.QUALIDADE_RESUMO.arquivo, config.QUALIDADE_RESUMO.aba
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("Tabela RES")
        ws.append(["título solto"])
        ws.append([None, "DATA", "OFICINA", "STATUS", "2QA", "PROD"])
        self.assertEqual(abas.resolver(wb, aba, arq), "Tabela RES")


class TestRegistroDeAbas(unittest.TestCase):
    def test_toda_aba_configurada_tem_assinatura(self):
        """Sem assinatura, uma aba renomeada volta a derrubar o pipeline."""
        usadas = {(f.arquivo, f.aba) for f in config.FONTES} | {
            (config.PRODUCAO.arquivo, config.PRODUCAO.aba),
            (config.ABSENTEISMO.arquivo, config.ABSENTEISMO.aba),
            (config.TREINO_EP.arquivo, config.TREINO_EP.aba),
            (config.TREINO_LIDERA.arquivo, config.TREINO_LIDERA.aba),
            (config.EFIC_JEANS.arquivo, config.EFIC_JEANS.aba),
            (config.EFIC_NAOJEANS.arquivo, config.EFIC_NAOJEANS.aba),
            (config.QUALIDADE_RESUMO.arquivo, config.QUALIDADE_RESUMO.aba),
            (config.QUALIDADE_DEFEITOS.arquivo, config.QUALIDADE_DEFEITOS.aba),
        }
        sem_assinatura = [
            chave for chave in sorted(usadas)
            if not config.aba_esperada(*chave).assinatura
        ]
        self.assertEqual(sem_assinatura, [])


if __name__ == "__main__":
    unittest.main()
