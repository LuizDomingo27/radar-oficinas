"""Leitura das colunas de VALOR das planilhas (Fase 2 — I/O isolado).

Cada função emite dicionários "crus" com chaves lógicas (nome, data, valores),
sem conhecer De-Para nem período — essa é a fronteira: a infra só sabe onde
estão as células (via ``config``); o serviço de consolidação enriquece depois.

As planilhas de eficiência são largas (pivotadas por semana); como são
pequenas, são lidas inteiras em memória para localizar os cabeçalhos de semana.
"""

from __future__ import annotations

import unicodedata
from pathlib import Path
from typing import Iterator

import openpyxl

from app_oficinas import config
from app_oficinas.errors import FonteInvalida, PlanilhaNaoEncontrada
from app_oficinas.infra import abas


def _abrir(caminho: Path):
    """Abre a planilha em modo somente-leitura, traduzindo falhas de I/O."""
    if not caminho.exists():
        raise PlanilhaNaoEncontrada(f"Planilha não encontrada: {caminho}")
    try:
        return openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    except Exception as exc:  # arquivo corrompido / não-xlsx
        raise PlanilhaNaoEncontrada(
            f"Falha ao abrir a planilha {caminho.name}: {exc}"
        ) from exc


def _aba(wb, nome_aba: str, arquivo: str):
    """Aba pedida, tolerando renomeação (ver ``infra.abas``)."""
    return abas.abrir_aba(wb, nome_aba, arquivo)


def _num(valor: object) -> float | None:
    """Converte uma célula em ``float`` de forma tolerante; ``None`` se inválida.

    O openpyxl (``data_only``) já devolve números nativos; o ramo de texto é só
    um fallback para células digitadas como string (aceita vírgula decimal).
    """
    if valor is None or isinstance(valor, bool):
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip()
    if not texto:
        return None
    try:
        return float(texto.replace(",", "."))
    except ValueError:
        return None


def _texto(valor: object) -> str | None:
    if valor is None:
        return None
    t = str(valor).strip()
    return t or None


def _celula(linha: tuple, indice: int):
    """Retorna a célula no índice, ou ``None`` se a linha for mais curta."""
    return linha[indice] if len(linha) > indice else None


# --------------------------------------------------------------------------- #
# Fontes com data real (uma linha = uma observação).                          #
# --------------------------------------------------------------------------- #

def ler_producao(base_dir: Path | None = None) -> Iterator[dict]:
    f = config.PRODUCAO
    wb = _abrir((base_dir or config.PLANILHAS_DIR) / f.arquivo)
    try:
        ws = _aba(wb, f.aba, f.arquivo)
        for linha in ws.iter_rows(min_row=f.primeira_linha, values_only=True):
            nome = _texto(_celula(linha, f.col_nome))
            if not nome:
                continue
            yield {
                "nome": nome,
                "mp": _texto(_celula(linha, f.col_mp)),
                "data": _celula(linha, f.col_data),
                "real_cortado": _num(_celula(linha, f.col_real_cortado)) or 0.0,
                "minutos": _num(_celula(linha, f.col_minutos)) or 0.0,
            }
    finally:
        wb.close()


def ler_absenteismo(base_dir: Path | None = None) -> Iterator[dict]:
    f = config.ABSENTEISMO
    wb = _abrir((base_dir or config.PLANILHAS_DIR) / f.arquivo)
    try:
        ws = _aba(wb, f.aba, f.arquivo)
        for linha in ws.iter_rows(min_row=f.primeira_linha, values_only=True):
            nome = _texto(_celula(linha, f.col_nome))
            if not nome:
                continue
            yield {
                "nome": nome,
                "frete": _texto(_celula(linha, f.col_frete)),
                "mp": _texto(_celula(linha, f.col_mp)),
                "data": _celula(linha, f.col_data),
                "semana": _texto(_celula(linha, f.col_semana)),
                "efetivos": _num(_celula(linha, f.col_efetivos)) or 0.0,
                "trabalhados": _num(_celula(linha, f.col_trabalhados)) or 0.0,
                "contratacao": _num(_celula(linha, f.col_contratacao)) or 0.0,
                "demissao": _num(_celula(linha, f.col_demissao)) or 0.0,
            }
    finally:
        wb.close()


def ler_treino_ep(base_dir: Path | None = None) -> Iterator[dict]:
    f = config.TREINO_EP
    wb = _abrir((base_dir or config.PLANILHAS_DIR) / f.arquivo)
    try:
        ws = _aba(wb, f.aba, f.arquivo)
        for linha in ws.iter_rows(min_row=f.primeira_linha, values_only=True):
            nome = _texto(_celula(linha, f.col_nome))
            if not nome:
                continue
            yield {
                "nome": nome,
                "modulo": _texto(_celula(linha, f.col_modulo)) or "(sem módulo)",
                "ch": _num(_celula(linha, f.col_ch)),
                "ciclo": _texto(_celula(linha, f.col_ciclo)),
            }
    finally:
        wb.close()


def ler_treino_lidera(base_dir: Path | None = None) -> Iterator[dict]:
    f = config.TREINO_LIDERA
    wb = _abrir((base_dir or config.PLANILHAS_DIR) / f.arquivo)
    try:
        ws = _aba(wb, f.aba, f.arquivo)
        for linha in ws.iter_rows(min_row=f.primeira_linha, values_only=True):
            nome = _texto(_celula(linha, f.col_nome))
            if not nome:
                continue
            yield {
                "nome": nome,
                "modulo": f.modulo,
                "data": _celula(linha, f.col_data),
                "polo": _texto(_celula(linha, f.col_polo)),
            }
    finally:
        wb.close()


# --------------------------------------------------------------------------- #
# Eficiência — planilhas largas despivotadas para formato longo.              #
# --------------------------------------------------------------------------- #

def ler_eficiencia_jeans(base_dir: Path | None = None) -> Iterator[dict]:
    """Eficiência oficial do JEANS: coluna "% 4WK" da aba ESTOQUE.

    Um valor por oficina (Méd. últimas 4 sem de entrega ÷ Cap Peças 100%), lido
    direto da planilha — é o número que a equipe usa. A coluna é achada pelo
    rótulo do cabeçalho, não por posição (a aba ganha colunas a cada carga).
    """
    f = config.EFIC_JEANS
    yield from _ler_efic_oficial(
        f.arquivo, f.aba, f.linha_cabecalho, f.primeira_linha, f.col_nome,
        mp_fixo=f.mp, base_dir=base_dir,
    )


def ler_eficiencia_naojeans(base_dir: Path | None = None) -> Iterator[dict]:
    """Eficiência oficial do NÃO-JEANS: coluna "MÉDIA 4W" da aba ESTOQUE OFICINAS.

    Mesma definição do JEANS; a MP (MALHA/POLO/TEAR) vem por linha.
    """
    f = config.EFIC_NAOJEANS
    yield from _ler_efic_oficial(
        f.arquivo, f.aba, f.linha_cabecalho, f.primeira_linha, f.col_nome,
        col_mp=f.col_mp, base_dir=base_dir,
    )


def _norm_rotulo(texto: str | None) -> str:
    """Normaliza um rótulo de cabeçalho: sem acento, maiúsculo, espaços colapsados."""
    if not texto:
        return ""
    semacento = "".join(
        c for c in unicodedata.normalize("NFD", texto) if not unicodedata.combining(c)
    )
    return " ".join(semacento.upper().split())


# Rótulos aceitos da coluna de % oficial, já normalizados (ver config).
_ROTULOS_EFIC = {_norm_rotulo(r) for r in config.ROTULOS_EFIC_OFICIAL}


def _achar_col_efic(cabecalho: tuple, arquivo: str, aba: str) -> int:
    """Índice da 1ª coluna cujo cabeçalho casa com a % oficial (peças).

    Levanta ``FonteInvalida`` se nenhuma casar — falha alto em vez de ler a
    coluna errada quando o layout da planilha muda.
    """
    for c in range(len(cabecalho)):
        if _norm_rotulo(_texto(_celula(cabecalho, c))) in _ROTULOS_EFIC:
            return c
    raise FonteInvalida(
        f"Coluna de eficiência não encontrada em {arquivo} / aba {aba}. "
        f"Procurei os rótulos {config.ROTULOS_EFIC_OFICIAL} na linha de "
        f"cabeçalho — confira se o nome da coluna mudou."
    )


def _ler_efic_oficial(arquivo, aba, linha_cabecalho, primeira_linha, col_nome,
                      mp_fixo=None, col_mp=None, base_dir=None) -> Iterator[dict]:
    """Lê a % de eficiência oficial (um valor por oficina) de uma aba de estoque."""
    wb = _abrir((base_dir or config.PLANILHAS_DIR) / arquivo)
    try:
        ws = _aba(wb, aba, arquivo)
        linhas = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if len(linhas) < linha_cabecalho:
        raise FonteInvalida(f"{arquivo} / aba {aba} não tem a linha de cabeçalho {linha_cabecalho}.")
    col_efic = _achar_col_efic(linhas[linha_cabecalho - 1], arquivo, aba)

    for linha in linhas[primeira_linha - 1:]:
        nome = _texto(_celula(linha, col_nome))
        if not nome:
            continue
        pct = _num(_celula(linha, col_efic))
        if pct is None:
            continue
        mp = mp_fixo if mp_fixo else _texto(_celula(linha, col_mp))
        yield _cel_efic(nome, mp, None, "efic_oficial", pct)


def _cel_efic(nome, mp, rotulo, tipo_valor, valor) -> dict:
    return {
        "nome": nome, "mp": mp, "rotulo_semana": rotulo,
        "tipo_valor": tipo_valor, "valor": valor, "ano": config.EFIC_JEANS.ano,
    }
