"""Leitura das abas de qualidade do "Indicador geral" (I/O isolado).

Emite dicionários crus, sem regra de negócio: ``ler_inspecoes`` (aba RESUMO)
e ``ler_defeitos`` (aba DEFEITOS). A agregação e as fórmulas ficam no serviço
``services/qualidade.py``. As abas são grandes (~41 mil e ~144 mil linhas),
então são lidas em modo streaming (``read_only`` + ``iter_rows``).

Otimização (cache): o "Indicador geral" tem ~24 MB e iterar as duas abas pelo
openpyxl custa ~22 s — abrir o arquivo em si é rápido (~0,4 s), o caro é
percorrer as 186 mil linhas de XML. Por isso, na primeira leitura as duas abas
são extraídas UMA vez para CSVs compactos (só as colunas usadas) em
``Planilhas/.cache_qualidade/``; as leituras seguintes leem os CSVs (~0,3 s). O
cache é validado pelo **hash do arquivo de origem**: enquanto o "Indicador
geral" não mudar, os uploads seguintes reaproveitam o cache (a maioria dos
uploads mexe só em outras planilhas). Se o cache não puder ser escrito, cai para
a leitura direta do Excel — correta, só não otimizada.
"""
from __future__ import annotations

import csv
import datetime as _dt
import hashlib
from pathlib import Path
from typing import Iterator

import openpyxl

from app_oficinas import config
from app_oficinas.errors import PlanilhaNaoEncontrada
from app_oficinas.infra import abas

_EPOCA_EXCEL = _dt.datetime(1899, 12, 30)

# Subpasta do cache (dentro de Planilhas/, que é ignorada pelo git).
_CACHE_SUBDIR = ".cache_qualidade"
# Versão do formato do cache: some ao mudar as colunas extraídas ou a lógica de
# parsing (config de coluna / _ano_mes). Bump = invalida caches antigos.
_CACHE_VERSAO = 1

_COLS_RESUMO = ("oficina", "status", "dois_qa", "prod", "ano", "mes")
_COLS_DEFEITOS = ("defeito", "tipo", "setor", "ano", "mes", "qntd")


def _abrir(caminho: Path):
    if not caminho.exists():
        raise PlanilhaNaoEncontrada(f"Planilha não encontrada: {caminho}")
    try:
        return openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    except Exception as exc:  # arquivo corrompido / não-xlsx
        raise PlanilhaNaoEncontrada(
            f"Falha ao abrir a planilha {caminho.name}: {exc}"
        ) from exc


def _aba(wb, nome_aba: str, arquivo: str):
    """Aba pedida, tolerando renomeação (ver ``infra.abas``)."""
    return abas.abrir_aba(wb, nome_aba, arquivo)


def _cel(row: tuple, indice: int):
    """Célula no índice, ou ``None`` se a linha for mais curta (evita IndexError)."""
    return row[indice] if len(row) > indice else None


def _texto(valor) -> str:
    return "" if valor is None else str(valor).replace("\t", " ").strip()


def _int(valor) -> int:
    try:
        return int(round(float(valor)))
    except (TypeError, ValueError):
        return 0


def _float(valor) -> float:
    if isinstance(valor, bool) or valor is None:
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    try:
        return float(str(valor).strip().replace(",", "."))
    except ValueError:
        return 0.0


def _ano_mes(valor) -> tuple[int | None, int | None]:
    """(ano, mês) a partir de um datetime ou serial Excel; (None, None) se vazio."""
    if isinstance(valor, (_dt.datetime, _dt.date)):
        return valor.year, valor.month
    if isinstance(valor, (int, float)) and not isinstance(valor, bool):
        d = _EPOCA_EXCEL + _dt.timedelta(days=int(valor))
        return d.year, d.month
    return None, None


# --------------------------------------------------------------------------- #
# Extração crua das abas (uma passada de openpyxl por aba).                    #
# --------------------------------------------------------------------------- #

def _extrair_inspecoes(wb) -> Iterator[dict]:
    """Cada inspeção da aba RESUMO: {oficina, status, dois_qa, prod, ano, mes}."""
    fonte = config.QUALIDADE_RESUMO
    ws = _aba(wb, fonte.aba, fonte.arquivo)
    for row in ws.iter_rows(min_row=fonte.primeira_linha, values_only=True):
        oficina = _texto(_cel(row, fonte.col_oficina))
        if not oficina:
            continue
        ano, mes = _ano_mes(_cel(row, fonte.col_data3))
        if ano is None:
            continue
        yield {
            "oficina": oficina,
            "status": _texto(_cel(row, fonte.col_status)).upper(),
            "dois_qa": _float(_cel(row, fonte.col_2qa)),
            "prod": _float(_cel(row, fonte.col_prod)),
            "ano": ano,
            "mes": mes,
        }


def _extrair_defeitos(wb) -> Iterator[dict]:
    """Cada defeito da aba DEFEITOS: {defeito, tipo, setor, ano, mes, qntd}."""
    fonte = config.QUALIDADE_DEFEITOS
    ws = _aba(wb, fonte.aba, fonte.arquivo)
    for row in ws.iter_rows(min_row=fonte.primeira_linha, values_only=True):
        defeito = _texto(_cel(row, fonte.col_defeito))
        if not defeito:
            continue
        yield {
            "defeito": defeito,
            "tipo": _texto(_cel(row, fonte.col_tipo)).upper(),
            "setor": _texto(_cel(row, fonte.col_setor)).upper(),
            "ano": _int(_cel(row, fonte.col_ano)) or None,
            "mes": _int(_cel(row, fonte.col_mes)) or None,
            "qntd": _float(_cel(row, fonte.col_qntd)),
        }


# --------------------------------------------------------------------------- #
# Cache CSV, validado pelo hash do "Indicador geral".                         #
# --------------------------------------------------------------------------- #

def _hash_arquivo(caminho: Path) -> str:
    """SHA-1 do conteúdo do arquivo (identifica mudança de forma robusta)."""
    h = hashlib.sha1()
    with open(caminho, "rb") as f:
        for bloco in iter(lambda: f.read(1 << 20), b""):
            h.update(bloco)
    return h.hexdigest()


def _chave_fonte(src: Path) -> str:
    """Chave de validade do cache: versão + tamanho + hash do arquivo de origem."""
    return f"{_CACHE_VERSAO}:{src.stat().st_size}:{_hash_arquivo(src)}"


def _escrever_csv(caminho: Path, colunas: tuple[str, ...], registros: Iterator[dict]) -> int:
    n = 0
    with open(caminho, "w", newline="", encoding="utf-8") as f:
        escritor = csv.writer(f)
        escritor.writerow(colunas)
        for r in registros:
            escritor.writerow(["" if r[c] is None else r[c] for c in colunas])
            n += 1
    return n


def garantir_cache(base_dir: Path | None = None, forcar: bool = False) -> tuple[Path, Path]:
    """Garante os CSVs de cache atualizados e devolve ``(resumo_csv, defeitos_csv)``.

    Recria o cache quando ele não existe, quando ``forcar`` é ``True`` ou quando o
    "Indicador geral" mudou (hash diferente). Caso contrário, reaproveita.

    Raises:
        PlanilhaNaoEncontrada: se o "Indicador geral" não existir/for ilegível.
    """
    base = base_dir or config.PLANILHAS_DIR
    src = base / config.QUALIDADE_RESUMO.arquivo
    if not src.exists():
        raise PlanilhaNaoEncontrada(f"Planilha não encontrada: {src}")

    cdir = base / _CACHE_SUBDIR
    resumo_csv = cdir / "resumo.csv"
    defeitos_csv = cdir / "defeitos.csv"
    chave_f = cdir / "fonte.chave"
    chave = _chave_fonte(src)

    if (not forcar and resumo_csv.exists() and defeitos_csv.exists()
            and chave_f.exists() and chave_f.read_text(encoding="utf-8") == chave):
        return resumo_csv, defeitos_csv

    cdir.mkdir(parents=True, exist_ok=True)
    wb = _abrir(src)
    try:
        _escrever_csv(resumo_csv, _COLS_RESUMO, _extrair_inspecoes(wb))
        _escrever_csv(defeitos_csv, _COLS_DEFEITOS, _extrair_defeitos(wb))
    finally:
        wb.close()
    chave_f.write_text(chave, encoding="utf-8")
    return resumo_csv, defeitos_csv


def _cache_ou_none(base_dir: Path | None) -> tuple[Path, Path] | None:
    """Cache pronto, ou ``None`` se não deu para usá-lo (cai na leitura direta).

    A ausência do arquivo de origem NÃO é tratada aqui: deixa a leitura direta
    (``_abrir``) levantar ``PlanilhaNaoEncontrada`` com a mensagem esperada.
    """
    base = base_dir or config.PLANILHAS_DIR
    if not (base / config.QUALIDADE_RESUMO.arquivo).exists():
        return None
    try:
        return garantir_cache(base_dir)
    except Exception:
        return None  # cache indisponível (ex.: disco só-leitura): leitura direta


# --------------------------------------------------------------------------- #
# API pública: lê do cache quando disponível; senão, direto do Excel.         #
# --------------------------------------------------------------------------- #

def ler_inspecoes(base_dir: Path | None = None) -> Iterator[dict]:
    """Cada inspeção da aba RESUMO: {oficina, status, dois_qa, prod, ano, mes}."""
    cache = _cache_ou_none(base_dir)
    if cache is not None:
        with open(cache[0], newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                yield {
                    "oficina": row["oficina"],
                    "status": row["status"],
                    "dois_qa": _float(row["dois_qa"]),
                    "prod": _float(row["prod"]),
                    "ano": int(row["ano"]),
                    "mes": int(row["mes"]),
                }
        return
    wb = _abrir((base_dir or config.PLANILHAS_DIR) / config.QUALIDADE_RESUMO.arquivo)
    try:
        yield from _extrair_inspecoes(wb)
    finally:
        wb.close()


def ler_defeitos(base_dir: Path | None = None) -> Iterator[dict]:
    """Cada defeito da aba DEFEITOS: {defeito, tipo, setor, ano, mes, qntd}."""
    cache = _cache_ou_none(base_dir)
    if cache is not None:
        with open(cache[1], newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                yield {
                    "defeito": row["defeito"],
                    "tipo": row["tipo"],
                    "setor": row["setor"],
                    "ano": int(row["ano"]) if row["ano"] else None,
                    "mes": int(row["mes"]) if row["mes"] else None,
                    "qntd": _float(row["qntd"]),
                }
        return
    wb = _abrir((base_dir or config.PLANILHAS_DIR) / config.QUALIDADE_DEFEITOS.arquivo)
    try:
        yield from _extrair_defeitos(wb)
    finally:
        wb.close()
